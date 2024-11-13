import os
import re
from time import strftime
import time

import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule


FLAT_CATEGORIES = [
    "1+kk",
    "1+1",
    "2+kk",
    "2+1",
    "3+kk",
    "3+1",
    "4+kk",
    "4+1",
    "5+kk",
    "5+1",
    "6 pokojů a více",
    "Atypický",
    "Pokoj",
]

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

FOLDER_NAME = "subory"

# Store all scraped properties
scraped_properties = []


def setup_webdriver() -> webdriver.Chrome:
    """Set up headless Chrome for Selenium."""
    options = Options()
    options.add_argument("--headless")
    driver = webdriver.Chrome(options=options)
    return driver


def generate_file_name() -> str:
    unique_values = df["category"].unique()
    file_name = ""

    for value in unique_values:
        file_name += f"{value}_"
    current_time = strftime("%d_%m_%Y_%H_%M")
    return f"{file_name+current_time}.xlsx"


driver = setup_webdriver()
wait = WebDriverWait(driver, 10)
url = input("vloz url: ")

# Inject script to handle Shadow DOM
driver.execute_cdp_cmd(
    "Page.addScriptToEvaluateOnNewDocument",
    {
        "source": """
Element.prototype._attachShadow = Element.prototype.attachShadow;
Element.prototype.attachShadow = function () {
    return this._attachShadow( { mode: "open" } );
};
"""
    },
)

driver.get(url)


def click_agree_with_ads_button() -> None:
    """
    Clicks the 'Agree with Ads' button inside a shadow DOM and waits for the URL to change,
    """
    time.sleep(2)
    closed_shadow_host = wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, ".szn-cmp-dialog-container"))
    )
    shadow_root = driver.execute_script(
        "return arguments[0].shadowRoot", closed_shadow_host
    )
    button = shadow_root.find_element(
        By.CSS_SELECTOR, "button[data-testid='cw-button-agree-with-ads']"
    )
    button.click()

    current_url = driver.current_url

    wait.until(lambda driver: driver.current_url != current_url)


click_agree_with_ads_button()


def scroll_page() -> None:
    """
    Scrolls the page to load more content. It simulates pressing the PAGE_DOWN key
    three times to ensure the page is fully scrolled.
    """
    for _ in range(3):
        ActionChains(driver).send_keys(Keys.PAGE_DOWN).perform()


def click_show_more() -> bool:
    """
    Attempts to click the 'Show More' button on the page. If the button is not found,
    the function returns False indicating that there are no more pages to load.

    Returns:
        bool: True if the 'Show More' button was found and clicked, False otherwise.
    """
    try:
        show_more_button = driver.find_element(
            By.CSS_SELECTOR, 'a[data-e2e="show-more-btn"]'
        )
        driver.execute_script("arguments[0].click();", show_more_button)
        return True
    except:
        return False


print("extrahujem data...")

while True:
    scroll_page()
    # If the 'Show More' button is not found, it means there are no more pages to load, so it starts scraping.
    if not click_show_more():
        properties_list = wait.until(
            EC.visibility_of_element_located(
                (By.CSS_SELECTOR, 'ul[data-e2e="estates-list"]')
            )
        )
        properties = properties_list.find_elements(By.TAG_NAME, "li")
        for item in properties:
            if item.get_attribute("id").startswith(
                "estate-list-item-"
            ) or item.get_attribute("id").startswith("region-tip-item-"):
                property_url = item.find_element(By.TAG_NAME, "a").get_attribute("href")
                property_id = property_url.split("/")[-1]
                property_info = item.find_elements(By.TAG_NAME, "p")
                property_description = property_info[0].text
                property_size = int(property_description.split(" ")[-2])
                property_address = property_info[1].text
                property_value = property_info[2].text
                property_category = re.search(
                    r"\b(?:"
                    + "|".join(re.escape(cat) for cat in FLAT_CATEGORIES)
                    + r")\b",
                    property_description,
                ).group()

                property_obj = {
                    "id": property_id,
                    "description": property_description,
                    "category": property_category,
                    "size": property_size,
                    "address": property_address,
                    "value": property_value,
                    "url": property_url,
                }
                if property_obj not in scraped_properties:
                    scraped_properties.append(property_obj)
        driver.quit()
        break

df = pd.DataFrame(scraped_properties)
df = df.sort_values(by=["size"])


def convert_rent_value(value):
    if "Kč" in value:
        return int(
            value.split("Kč")[0].replace(" ", "").replace("měsíc", "").replace("/", "")
        )
    else:
        return None


df["rent_price"] = df["value"].apply(convert_rent_value)

interval_size = 10

# Find the minimum and maximum sizes in the dataset
min_size = df["size"].min()

df["size_interval"] = None

for index, row in df.iterrows():
    # Calculate the size interval for the current row
    size_interval = (
        int((((row["size"] - min_size) // 10) * 10) + min_size),
        int((((row["size"] - min_size) // 10) * 10) + min_size + 10),
    )

    df.at[index, "size_interval"] = size_interval

df = df.sort_values(by=["category", "size"])

file_name = generate_file_name()
file_path = os.path.join(FOLDER_NAME, file_name)

# Start writing to an Excel file
with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
    for category, group in df.groupby("category"):
        # Create a new sheet for each category
        sheet_name = category
        workbook = writer.book
        workbook.create_sheet(title=sheet_name)
        sheet = workbook[sheet_name]

        header_font = Font(bold=True, color="000000")
        header_fill = PatternFill(
            start_color="FF6600", end_color="FF6600", fill_type="solid"
        )

        row_offset = 1
        for interval, interval_group in group.groupby("size_interval"):
            if not interval_group.empty:
                interval_title = f"{interval[0]} - {interval[1]} (m²)"

                # Title cell styling
                title_cell = sheet.cell(row=row_offset, column=1, value=interval_title)
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                title_cell.font = Font(bold=True)
                title_cell.fill = PatternFill(
                    start_color="FF6600", end_color="FF6600", fill_type="solid"
                )
                title_cell.border = THIN_BORDER
                sheet.merge_cells(
                    start_row=row_offset,
                    start_column=1,
                    end_row=row_offset,
                    end_column=4,
                )

                # Add table headers
                headers = ["flat_size (m²)", "address", "price (Kč)", "url"]
                for col, header in enumerate(headers, start=1):
                    header_cell = sheet.cell(
                        row=row_offset + 1, column=col, value=header
                    )
                    header_cell.font = header_font
                    header_cell.fill = header_fill
                    header_cell.alignment = Alignment(
                        horizontal="center", vertical="center"
                    )
                    header_cell.border = THIN_BORDER

                # Write data rows for the current interval
                for idx, row in enumerate(
                    interval_group.itertuples(), start=row_offset + 2
                ):
                    flat_size, address, price, url = (
                        row.size,
                        row.address,
                        row.rent_price,
                        row.url,
                    )

                    # Write each value to the corresponding cell in the row
                    for col_index, value in enumerate(
                        [flat_size, address, price, url], start=1
                    ):
                        cell = sheet.cell(row=idx, column=col_index, value=value)
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                        if col_index % 2 == 0:
                            cell.alignment = Alignment(
                                horizontal="left", vertical="center"
                            )
                        cell.border = THIN_BORDER

                # Calculate statistics for the 'rent_price' column
                start_data_row = row_offset + 2
                end_data_row = sheet.max_row

                avg_price = f"AVERAGE(C{start_data_row}:C{end_data_row})"
                max_price = f"MAX(C{start_data_row}:C{end_data_row})"
                min_price = f"MIN(C{start_data_row}:C{end_data_row})"

                # Write statistics below each interval table
                row_offset = sheet.max_row + 2
                sheet.cell(row=row_offset, column=1, value="AVG price:").border = (
                    THIN_BORDER
                )
                sheet.cell(row=row_offset, column=2, value=f"={avg_price}").border = (
                    THIN_BORDER
                )

                avg_price_cell = f"$B${sheet.max_row}"

                # Apply conditional formatting for prices above and below average
                above_avg_fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )  # red accent 2, lighter 40%
                below_equal_avg_fill = PatternFill(
                    start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
                )  # olive green, accent 3, lighter 60%

                # Define the range where conditional formatting will apply (price column in the specified row range)
                price_range = f"C{start_data_row}:C{end_data_row}"

                # Conditional formatting formulas
                above_avg_formula = f"C{start_data_row} > {avg_price_cell}"  # Formula to check if cell value is above average
                below_equal_avg_formula = f"C{start_data_row} <= {avg_price_cell}"  # Formula to check if cell value is below or equal to average

                # Apply the conditional formatting rules
                sheet.conditional_formatting.add(
                    price_range,
                    FormulaRule(formula=[above_avg_formula], fill=above_avg_fill),
                )
                sheet.conditional_formatting.add(
                    price_range,
                    FormulaRule(
                        formula=[below_equal_avg_formula], fill=below_equal_avg_fill
                    ),
                )
                row_offset += 1
                sheet.cell(row=row_offset, column=1, value="MAX price:").border = (
                    THIN_BORDER
                )
                sheet.cell(row=row_offset, column=2, value=f"={max_price}").border = (
                    THIN_BORDER
                )
                row_offset += 1
                sheet.cell(row=row_offset, column=1, value="MIN price:").border = (
                    THIN_BORDER
                )
                sheet.cell(row=row_offset, column=2, value=f"={min_price}").border = (
                    THIN_BORDER
                )

                row_offset += 2

        # Adjust column widths for columns A, B, C, and D
        for col in range(1, 5):  # Columns A to D
            max_length = 0
            col_letter = get_column_letter(col)
            for row in sheet.iter_rows(
                min_row=1, max_row=sheet.max_row, min_col=col, max_col=col
            ):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            # Set the column width based on the maximum content length plus some padding
            sheet.column_dimensions[col_letter].width = max_length + 2

print(f"Subor {file_name} je ulozeny v priecinku {FOLDER_NAME}")
